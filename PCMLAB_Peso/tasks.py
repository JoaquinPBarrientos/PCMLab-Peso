from robocorp.tasks import task
import pandas as pd 

import tkinter as tk
from tkinter import filedialog

import math
import pandas as pd 

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

NAME_KEY = 'Nombre Destinatario'
WEIGHT_KEY = 'Cantidad UMV'
TIME_KEY = 'Fec Guía Despacho'

WEIGHT_WORD = 'Peso (Kg)'
TIME_WORD = 'Fecha de Despacho'


@task
def PCMLab_monthly_report():
    """ Program for generating a monthly report of the PCMLab,
        this program will generate an Excel file with the daily weights
        as well as the monthly weight for each distributor"""
    
    file_path = ask_file()
    RESULTS_PATH = guardar_archivo()
    input = excel_preprocessing(file_path)
    report_by_day(input, RESULTS_PATH)
    report_by_enterprise(input,RESULTS_PATH)
    


def ask_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path

def excel_preprocessing(file_path):
    """
    Preprocesses an Excel file by reading it, selecting specific columns, and converting the time column to datetime.

    Args:
        file_path (str): The path to the Excel file.

    Returns:
        pandas.DataFrame: The preprocessed DataFrame containing the selected columns and converted time column.
    """
    file = pd.read_excel(file_path)
    columns = [NAME_KEY, WEIGHT_KEY, TIME_KEY]
    file = file[columns]
    file[TIME_KEY] = pd.to_datetime(file[TIME_KEY])
    return file

def report_by_day(file, RESULTS_PATH):
    """
    Generate a report by day based on the given file.

    Args:
        file (DataFrame): The input file containing weight data.

    Returns:
        None
    """

    try:
        results = pd.DataFrame()

        # Get the rows, that are dates, and the are sorted
        rows = file[TIME_KEY].dt.strftime('%d-%m-%Y').values

        rows = [x for x in rows if x is not None and x != 'nan']
        rows = (list(set(rows)))
        rows = sorted(rows, key=lambda x: int(str(x).split('-')[0]))
    except Exception as e:
        print(f"Error processing date column: {e}")
        
    # Create a DataFrame with the dates as index
    month = pd.DataFrame()
    month.index = rows
    month.index.name = TIME_WORD
    
    # Reindex the DataFrame
    month = month.reindex(file[TIME_KEY].dt.strftime('%d-%m-%Y').unique())
    
    # Get a list of the total weights for each day
    weights_by_day = file.groupby(file[TIME_KEY].dt.strftime('%d-%m-%Y'))[WEIGHT_KEY].sum()
    
    # Assign the weights to the DataFrame
    month[WEIGHT_WORD] = weights_by_day.reindex(month.index)
    
    # Gets the total weight for the month
    total_weight = month[WEIGHT_WORD].sum()
    month.loc['Total'] = total_weight
    results = pd.concat([results, month])
    results.to_excel(RESULTS_PATH)
    
    # Delete 
    del results
    del month
    # Call minor fixes
    minor_fixes(RESULTS_PATH)


def minor_fixes(RESULTS_PATH):
    """
    Performs minor fixes on the workbook and sheet.

    This function opens the workbook, inserts a title to the table, and saves the changes.

    Parameters:
    None

    Returns:
    None
    """
    book = load_workbook(RESULTS_PATH)
    sheet = book.active
    sheet.insert_rows(0)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    sheet.cell(row=1, column=1).value = 'Reporte por día'
    book.save(RESULTS_PATH)

def report_by_enterprise(file,RESULTS_PATH):
    """
    Generates a report by enterprise based on the given file.

    Args:
        file (pandas.DataFrame): The input file containing enterprise data.

    Returns:
        None
    """

    df_enterprises = pd.DataFrame()
    enterprises = file[NAME_KEY].values
    enterprises = list(set(enterprises))
    enterprises = [x for x in enterprises if isinstance(x, str)]
    df_enterprises.index = enterprises
    df_enterprises.index.name = 'Empresa'
    df_enterprises['Peso Total'] = 0

    print(enterprises)
    # Obteins the total weight for each enterprise 
    for enterprise in enterprises:
        df_enterprises.loc[enterprise] = file.loc[file[NAME_KEY] == enterprise][WEIGHT_KEY].sum()

    df_enterprises.loc['Total'] = df_enterprises['Peso Total'].sum()
    df_results = pd.read_excel(RESULTS_PATH, index_col=0)

    # Unify both dataframes and write them to the Excel file
    with pd.ExcelWriter(RESULTS_PATH, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=0, index=True)
        df_enterprises.to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=len(df_results.columns) + 2, index=True)

def guardar_archivo():
    # Abre el diálogo para guardar archivo
    archivo = filedialog.asksaveasfile(defaultextension=".xlsx",
                                       filetypes=[("Todos los archivos", "*.*")])
    return archivo.name